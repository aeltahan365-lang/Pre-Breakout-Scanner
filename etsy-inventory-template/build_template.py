# -*- coding: utf-8 -*-
"""
Momentum Mind Store - "Inventory Manager PRO" v2
A no-VBA, mobile-app-styled Inventory / POS / Invoicing workbook.

Core idea: THE SALE IS LOGGED THE MOMENT IT IS SCANNED.
Both logs (SKU-level and header-level) are live formula views of the entry
screen, so there is no "post" step and nothing can go unlogged.

Run:  python3 build_template.py
"""
import datetime as _dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter as CL

# ── DESIGN TOKENS: "Momentum Midnight" ───────────────────────────────────────
INK, SLATE, MUTED, BORDER = "0F172A", "334155", "64748B", "E2E8F0"
BG, CARD, WHITE = "F1F5F9", "FFFFFF", "FFFFFF"
INDIGO, INDIGO_DK, INDIGO_LT = "4F46E5", "3730A3", "EEF2FF"
GREEN, GREEN_BG = "047857", "ECFDF5"
RED, RED_BG = "B91C1C", "FEE2E2"
AMBER, AMBER_BG = "B45309", "FFFBEB"
WA_GREEN, INPUT_BG, INPUT_BRD = "25D366", "FFF9DB", "F59E0B"
FONT = "Segoe UI"

def fnt(sz=11, b=False, c=INK, i=False, u=None):
    return Font(name=FONT, size=sz, bold=b, color=c, italic=i, underline=u)
def fill(h): return PatternFill("solid", fgColor=h)
def al(h="left", v="center", wrap=False, ind=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=ind)

HAIR = Side(style="thin", color=BORDER)
BOX = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)
def gold():
    s = Side(style="medium", color=INPUT_BRD)
    return Border(left=s, right=s, top=s, bottom=s)

MONEY = '$#,##0.00;[Red]($#,##0.00);"-"'
INT = '#,##0;[Red](#,##0);"-"'
PCT = '0.0%'
DATE = 'dd-mmm-yyyy'

# ── NAV: six primary screens ────────────────────────────────────────────────
NAV = {
    "home":  ("🏠 Home",      "Dashboard"),
    "items": ("📦 Items",     "Item Master"),
    "sell":  ("🛒 New Sale",  "Log Sale"),
    "stock": ("📥 Stock In",  "Stock In"),
    "inv":   ("🧾 Invoice",   "Invoice Generator"),
    "logs":  ("📊 Logs",      "Sales Log"),
}

def shell(ws, last_col, slots, freeze, tab, rows=340):
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False
    ws.sheet_properties.tabColor = tab
    ws.freeze_panes = freeze
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 6
    for r in range(1, rows):
        for c in range(1, last_col + 1):
            ws.cell(r, c).fill = fill(BG)
    for c in range(1, last_col + 1):
        ws.cell(1, c).fill = fill(INK)
        ws.cell(2, c).fill = fill(INDIGO)
    for rng, key in slots:
        label, target = NAV[key]
        if ":" in rng: ws.merge_cells(rng)
        cell = ws[rng.split(":")[0]]
        cell.value = '=HYPERLINK("#\'{}\'!A1","{}")'.format(target, label)
        cell.font = fnt(11, True, WHITE)
        cell.fill = fill(INK)
        cell.alignment = al("center", "center")

def widths(ws, spec):
    for k, v in spec.items(): ws.column_dimensions[k].width = v
def heights(ws, a, b, h):
    for r in range(a, b + 1): ws.row_dimensions[r].height = h
def put(ws, ref, val, font=None, bg=None, a=None, nf=None, bd=None):
    c = ws[ref]; c.value = val
    if font: c.font = font
    if bg: c.fill = fill(bg)
    if a: c.alignment = a
    if nf: c.number_format = nf
    if bd: c.border = bd
    return c
def title(ws, c1, c2, kicker, head, sub):
    for r, h in ((4, 20), (5, 34), (6, 20)):
        ws.merge_cells(f"{c1}{r}:{c2}{r}"); ws.row_dimensions[r].height = h
    put(ws, f"{c1}4", kicker, fnt(9, True, INDIGO), BG, al("left", "bottom", ind=1))
    put(ws, f"{c1}5", head, fnt(22, True, INK), BG, al("left", "center", ind=1))
    put(ws, f"{c1}6", sub, fnt(10, False, MUTED), BG, al("left", "top", ind=1))
def header_row(ws, row, labels, height=36):
    ws.row_dimensions[row].height = height
    for i, h in enumerate(labels, 1):
        put(ws, f"{CL(i)}{row}", h, fnt(10, True, WHITE), INK, al("center", "center", True), bd=BOX)

def XL(look, arr, ret, nf='""'):
    return f'_xlfn.XLOOKUP({look},{arr},{ret},{nf})'

# ── SHEET RANGES (single source of truth) ───────────────────────────────────
IM, LS, SI, SL = "'Item Master'", "'Log Sale'", "'Stock In'", "'Sales Log'"
IM_ROWS, LS_ROWS, SI_ROWS, SL_ROWS = (4, 53), (9, 308), (8, 207), (8, 107)

def R(sheet, col, rows): return f"{sheet}!${col}${rows[0]}:${col}${rows[1]}"
# Item Master: A SKU · B Barcode · C Name · D Category · E Cost · F Price
#              G Max · H Reorder · I Supplier · J Phone · K Email
i_sku, i_bar, i_name = R(IM,'A',IM_ROWS), R(IM,'B',IM_ROWS), R(IM,'C',IM_ROWS)
i_cost, i_price, i_max = R(IM,'E',IM_ROWS), R(IM,'F',IM_ROWS), R(IM,'G',IM_ROWS)
i_rpt, i_supp = R(IM,'H',IM_ROWS), R(IM,'I',IM_ROWS)
i_phone, i_mail = R(IM,'J',IM_ROWS), R(IM,'K',IM_ROWS)
# Log Sale: A 🆕 · B Date · C Invoice · D Scan · E Product · F Qty · G Price
#           H Total · I _SKU · J _Key · K _Qty
l_new, l_date, l_inv, l_scan = R(LS,'A',LS_ROWS), R(LS,'B',LS_ROWS), R(LS,'C',LS_ROWS), R(LS,'D',LS_ROWS)
l_tot, l_sku, l_key, l_qty = R(LS,'H',LS_ROWS), R(LS,'I',LS_ROWS), R(LS,'J',LS_ROWS), R(LS,'K',LS_ROWS)
# Stock In: A Date · B Ref · C Scan · D Type · E Product · F Qty · G Cost · H Total · I _SKU
s_type, s_qty, s_sku = R(SI,'D',SI_ROWS), R(SI,'F',SI_ROWS), R(SI,'I',SI_ROWS)
# Sales Log: A Inv · B Date · C Cust · D Phone · E Email · F Items · G Qty
#            H Subtotal · I Tax · J Amount · K _pick
g_inv, g_date, g_cust = R(SL,'A',SL_ROWS), R(SL,'B',SL_ROWS), R(SL,'C',SL_ROWS)
g_phone, g_mail, g_amt = R(SL,'D',SL_ROWS), R(SL,'E',SL_ROWS), R(SL,'J',SL_ROWS)

PREFIX = "Settings!$C$14"
TAXRATE = "Settings!$C$12"

wb = openpyxl.Workbook(); wb.remove(wb.active)

# ════════════════════════════════════════════════════════════════════════════
# 1 · START HERE  — every step card is itself the link to that screen
# ════════════════════════════════════════════════════════════════════════════
sh = wb.create_sheet("Start Here")
widths(sh, {"A":3,"B":18,"C":18,"D":18,"E":18,"F":18,"G":18,"H":3})
shell(sh, 8, [("A1:B1","home"),("C1","items"),("D1","sell"),
              ("E1","stock"),("F1","inv"),("G1:H1","logs")], "A3", INDIGO_DK)
for r,h in ((4,20),(5,38),(6,22)):
    sh.merge_cells(f"B{r}:G{r}"); sh.row_dimensions[r].height = h
put(sh,"B4","MOMENTUM MIND STORE",fnt(9,True,INDIGO),BG,al("left","bottom"))
put(sh,"B5","📦 Inventory Manager PRO",fnt(24,True,INK),BG,al("left","center"))
put(sh,"B6","Scan. Sell. Send. Your whole shop in one file — tap a card to jump straight in.",
    fnt(10,False,MUTED),BG,al("left","top"))

STEPS = [
    ("Item Master","📦  STEP 1 — Add your products",
     "One row per item. Give each a unique SKU and, if you use a scanner, its barcode. "
     "Cost, Price, Max Stock and supplier contact all live here — every other screen reads from it."),
    ("Settings","⚙️  STEP 2 — Set your business details",
     "Shop name, contact details, tax rate and invoice prefix. Every invoice and every "
     "auto-drafted supplier message reads from this one screen."),
    ("Log Sale","🛒  STEP 3 — Sell (this IS your sales log)",
     "Tick 🆕 to start a new sale, scan or pick the item, type the quantity. That is the whole "
     "sale. The invoice number, description, price and line total fill themselves — and the sale "
     "is logged the instant you enter it. There is no save button because nothing needs saving."),
    ("Stock In","📥  STEP 4 — Receive stock & record breakages",
     "Deliveries from suppliers go here as IN. Damaged or lost items go here as SHRINKAGE. "
     "Both feed the live stock count on the Dashboard."),
    ("Invoice Generator","🧾  STEP 5 — Send the invoice",
     "The Invoice screen already shows your most recent sale. Tap 💬 to send it on WhatsApp or "
     "✉️ to email it — the whole invoice is written for you. For a PDF: File ▸ Export ▸ PDF."),
    ("Sales Log","📊  STEP 6 — Read your two live logs",
     "Log Sale is your line-by-line SKU log. Sales Log is the header view — one row per invoice "
     "with date, invoice number, quantity and amount. Both are automatic."),
    ("Dashboard","🏠  STEP 7 — Watch the Dashboard",
     "Live stock, today's takings, and low-stock alerts with one-tap WhatsApp and Email reorder "
     "buttons that draft the message to your supplier for you."),
]
row = 8
for target, head, body in STEPS:
    sh.merge_cells(f"B{row}:G{row}"); sh.merge_cells(f"B{row+1}:G{row+1}")
    put(sh,f"B{row}",f'=HYPERLINK("#\'{target}\'!A1","{head}   ›")',
        fnt(12,True,INDIGO_DK),CARD,al("left","center",ind=1))
    put(sh,f"B{row+1}",body,fnt(10,False,SLATE),CARD,al("left","top",True,1))
    for c in range(2,8):
        for rr in (row,row+1):
            sh.cell(rr,c).fill = fill(CARD); sh.cell(rr,c).border = BOX
    sh.row_dimensions[row].height = 32; sh.row_dimensions[row+1].height = 48
    row += 3

sh.merge_cells(f"B{row}:G{row}")
put(sh,f"B{row}","🎨 COLOUR KEY",fnt(11,True,INDIGO),BG,al("left","center")); row += 1
for lab, desc in [("🟡  Yellow","You type or scan here."),
                  ("⬜  White","Calculated automatically — do not overwrite."),
                  ("🟦  Dark bar","Six screens, one tap away, on every sheet."),
                  ("🟩  Green / 🟪 Indigo","A button that has switched itself on — tap it.")]:
    put(sh,f"B{row}",lab,fnt(10,True,INK),CARD,al("left","center",ind=1),bd=BOX)
    sh.merge_cells(f"C{row}:G{row}")
    put(sh,f"C{row}",desc,fnt(10,False,SLATE),CARD,al("left","center",ind=1))
    for c in range(2,8): sh.cell(row,c).fill = fill(CARD); sh.cell(row,c).border = BOX
    sh.row_dimensions[row].height = 30; row += 1

row += 1
sh.merge_cells(f"B{row}:G{row}")
put(sh,f"B{row}","💬 NEED HELP?",fnt(11,True,INDIGO),BG,al("left","center")); row += 1
for lab, f_ in [
    ("✉️  Support email",'=HYPERLINK("mailto:support@momentummindstore.com?subject=Inventory%20Manager%20PRO","support@momentummindstore.com")'),
    ("🛍️  Etsy shop",'=HYPERLINK("https://www.etsy.com/shop/MomentumMindStore","etsy.com/shop/MomentumMindStore")'),
    ("📖  Setup video",'=HYPERLINK("https://www.momentummindstore.com/inventory-pro-guide","Watch the 4-minute walkthrough")'),
    ("⚙️  Settings",'=HYPERLINK("#\'Settings\'!A1","Open your business details")'),
    ("⭐  Loved it?",'=HYPERLINK("https://www.etsy.com/your/purchases","Leave a review — it genuinely helps a small shop")')]:
    put(sh,f"B{row}",lab,fnt(10,True,INK),CARD,al("left","center",ind=1),bd=BOX)
    sh.merge_cells(f"C{row}:G{row}")
    put(sh,f"C{row}",f_,fnt(10,True,INDIGO),CARD,al("left","center",ind=1))
    for c in range(2,8): sh.cell(row,c).fill = fill(CARD); sh.cell(row,c).border = BOX
    sh.row_dimensions[row].height = 30; row += 1
row += 1
sh.merge_cells(f"B{row}:G{row}")
put(sh,f"B{row}","© Momentum Mind Store — single-shop licence. No resale or redistribution of this file.",
    fnt(9,False,MUTED,i=True),BG,al("left","center"))

# ════════════════════════════════════════════════════════════════════════════
# 2 · ITEM MASTER  — tbl_Items, now scanner-ready
# ════════════════════════════════════════════════════════════════════════════
im = wb.create_sheet("Item Master")
widths(im, {"A":12,"B":14,"C":30,"D":14,"E":11,"F":11,"G":11,"H":12,"I":20,"J":16,"K":24})
shell(im, 11, [("A1:B1","home"),("C1","items"),("D1:E1","sell"),
               ("F1:G1","stock"),("H1:I1","inv"),("J1:K1","logs")], "A4", INDIGO)
header_row(im, 3, ["SKU","📷 Barcode","📦 Product Name","Category","Cost","Price",
                   "Max Stock","Reorder Point","Supplier Name","📞 Supp. Phone","✉️ Supp. Email"], 38)
PRODUCTS = [
    ("LAV-001","5060001000017","🕯️ Lavender Soy Candle 220g","Home Fragrance",6.50,18.00,120,25,"Aurora Wax Co.","15551234567","orders@aurorawax.com"),
    ("CER-014","5060001000024","☕ Ceramic Mug — Sand","Drinkware",4.20,14.00,200,40,"Kiln & Clay Ltd","15559876543","sales@kilnclay.com"),
    ("TOT-007","5060001000031","👜 Canvas Tote Bag — Natural","Bags",3.80,12.50,150,30,"NorthLoom Textiles","15554567890","hello@northloom.com"),
    ("JRN-002","5060001000048","📓 Linen Journal A5","Stationery",5.10,16.00,100,20,"PaperFold Studio","15553216549","supply@paperfold.com"),
    ("SKN-021","5060001000055","🧴 Botanical Body Oil 100ml","Skincare",7.90,24.00,80,16,"Botanica Labs","15557778888","orders@botanicalabs.com"),
    ("TEA-009","5060001000062","🍵 Loose Leaf Tea 200g","Pantry",5.60,17.50,90,18,"Verde Leaf Imports","15552223333","buy@verdeleaf.com"),
]
heights(im, 4, 53, 32)
for r in range(4, 54):
    i = r - 4
    if i < len(PRODUCTS):
        for c, v in enumerate(PRODUCTS[i], 1): im.cell(r, c, v)
    for c in range(1, 12):
        cell = im.cell(r, c)
        cell.fill = fill(INPUT_BG); cell.border = BOX
        cell.font = fnt(10, c == 1, INK)
        cell.alignment = al("center" if c in (1,2,5,6,7,8) else "left","center",
                            ind=0 if c in (1,2,5,6,7,8) else 1)
        if c in (5,6): cell.number_format = MONEY
        if c in (7,8): cell.number_format = INT
        if c in (2,10): cell.number_format = "@"
t = Table(displayName="tbl_Items", ref="A3:K53")
t.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True,
                                  showColumnStripes=False, showFirstColumn=False, showLastColumn=False)
im.add_table(t)
im["A3"].comment = Comment(
    "SKU must be UNIQUE — it is the key linking every screen.\n\n"
    "📷 Barcode: scan straight into this cell with any USB or Bluetooth scanner "
    "(they type like a keyboard). The Sell and Stock In screens accept EITHER the "
    "SKU or the barcode.\n\n"
    "📞 Phone: digits only, country code first, no + or spaces (e.g. 15551234567).", "Momentum Mind Store")

# ════════════════════════════════════════════════════════════════════════════
# 3 · LOG SALE  — the till. Entering a line IS logging the sale.
# ════════════════════════════════════════════════════════════════════════════
LATEST = f'IFERROR(INDEX({g_inv},MAX(1,COUNTIF({g_inv},"?*"))),"")'
ls = wb.create_sheet("Log Sale")
widths(ls, {"A":6,"B":13,"C":15,"D":17,"E":30,"F":8,"G":11,"H":13,"I":12,"J":16,"K":8})
for c in "IJK": ls.column_dimensions[c].hidden = True
shell(ls, 11, [("A1:B1","home"),("C1","items"),("D1","sell"),
               ("E1","stock"),("F1:G1","inv"),("H1","logs")], "A9", GREEN)
title(ls,"A","H","POINT OF SALE","🛒 New Sale",
      "Tick 🆕 · scan or pick the item · type the quantity. That is the entire sale.")

ls.row_dimensions[7].height = 46
CARD_CELLS = [("A7:C7", f'="🧾  "&IF({LATEST}="","— no sales yet —",{LATEST})', None, INDIGO_LT, INDIGO_DK, 14),
              ("D7:E7", f'=IF({LATEST}="",0,COUNTIF({l_inv},{LATEST}))', INT, INDIGO_LT, INDIGO_DK, 14),
              ("F7:G7", f'=IF({LATEST}="",0,SUMIF({l_inv},{LATEST},{l_qty}))', INT, INDIGO_LT, INDIGO_DK, 14),
              ("H7",    f'=IF({LATEST}="",0,SUMIF({l_inv},{LATEST},{l_tot}))', MONEY, GREEN_BG, GREEN, 14)]
CARD_TIPS = ["Current sale", "Lines on this sale", "Units on this sale", "Running total"]
for (rng, f_, nf, bgc, txt, sz), tip in zip(CARD_CELLS, CARD_TIPS):
    if ":" in rng: ls.merge_cells(rng)
    a = rng.split(":")[0]
    put(ls, a, f_, fnt(sz, True, txt), bgc, al("center","center"), nf, BOX)
    ls[a].comment = Comment(tip, "Momentum Mind Store")
    end = rng.split(":")[-1]
    for c in range(openpyxl.utils.column_index_from_string(a[0]),
                   openpyxl.utils.column_index_from_string(end[0]) + 1):
        ls.cell(7, c).fill = fill(bgc); ls.cell(7, c).border = BOX

header_row(ls, 8, ["🆕","📅 Date","🧾 Invoice No","🔎 SKU / Scan","📦 Product",
                   "🔢 Qty","💲 Price","💵 Line Total","_SKU","_Key","_Qty"], 38)
A, B = LS_ROWS
heights(ls, A, B, 32)
SALES = [("x","2026-08-05","LAV-001",2),("","2026-08-05","CER-014",3),("","2026-08-05","JRN-002",1),
         ("x","2026-08-07","TOT-007",4),("","2026-08-07","TEA-009",2),
         ("x","2026-08-09","SKN-021",55),
         ("x","2026-08-10","LAV-001",88),
         ("x","2026-08-12","CER-014",12)]
for r in range(A, B + 1):
    i = r - A
    if i < len(SALES):
        nw, d, sku, q = SALES[i]
        if nw: ls.cell(r,1,nw)
        ls.cell(r,2,_dt.date(*[int(x) for x in d.split("-")]))
        ls.cell(r,4,sku); ls.cell(r,6,q)
    put(ls,f"C{r}",f'=IF($D{r}="","",IF(OR($A{r}="x",ROW()={A}),{PREFIX}&TEXT(1000+COUNTIF($A${A}:$A{r},"x")+IF($A${A}<>"x",1,0),"0000"),$C{r-1}))',
        fnt(11,True,INDIGO_DK),CARD,al("center","center"),bd=BOX)
    put(ls,f"I{r}",f'=IF($D{r}="","",IF(COUNTIF({i_sku},$D{r})>0,$D{r},{XL(f"$D{r}",i_bar,i_sku)}))',
        fnt(9,False,MUTED),CARD,al("center","center"),bd=BOX)
    put(ls,f"E{r}",f'=IF($D{r}="","",IF($I{r}="","⚠️ Unknown code",{XL(f"$I{r}",i_sku,i_name)}))',
        fnt(10,False,INK),CARD,al("left","center",ind=1),bd=BOX)
    put(ls,f"G{r}",f'=IF($I{r}="","",{XL(f"$I{r}",i_sku,i_price,"0")})',
        fnt(10,False,MUTED),CARD,al("center","center"),MONEY,BOX)
    put(ls,f"K{r}",f'=IF($I{r}="","",IF($F{r}="",1,$F{r}))',fnt(9,False,MUTED),CARD,al("center","center"),INT,BOX)
    put(ls,f"H{r}",f'=IF($I{r}="","",$K{r}*$G{r})',fnt(11,True,INK),CARD,al("center","center"),MONEY,BOX)
    put(ls,f"J{r}",f'=IF($I{r}="","",$C{r}&"|"&COUNTIFS($C${A}:$C{r},$C{r},$I${A}:$I{r},"?*"))',
        fnt(9,False,MUTED),CARD,al("center","center"),bd=BOX)
    for c,(a_,f_) in {1:("center",None),2:("center",DATE),4:("center",None),6:("center",INT)}.items():
        cell = ls.cell(r,c); cell.fill = fill(INPUT_BG); cell.border = BOX
        cell.font = fnt(11, c in (1,4), INK); cell.alignment = al(a_,"center")
        if f_: cell.number_format = f_

dv = DataValidation(type="list", formula1='"x"', allow_blank=True, showDropDown=False)
dv.prompt = "Tap and choose x to start a NEW sale. Leave blank to add another line to the sale above."
dv.promptTitle = "🆕 New sale?"
ls.add_data_validation(dv); dv.add(f"A{A}:A{B}")
dv2 = DataValidation(type="list", formula1="SKU_List", allow_blank=True,
                     showDropDown=False, showErrorMessage=False)
dv2.prompt = "Tap the arrow to pick, start typing to search, or scan a barcode straight in."
dv2.promptTitle = "🔎 Scan or select"
ls.add_data_validation(dv2); dv2.add(f"D{A}:D{B}")
dv3 = DataValidation(type="whole", operator="greaterThan", formula1="0", allow_blank=True)
dv3.prompt = "Leave blank for 1."; dv3.promptTitle = "🔢 Quantity"
dv3.error = "Quantity must be a whole number above zero."; dv3.errorTitle = "Invalid quantity"
ls.add_data_validation(dv3); dv3.add(f"F{A}:F{B}")

ls.conditional_formatting.add(f"E{A}:E{B}", FormulaRule(
    formula=[f'$E{A}="⚠️ Unknown code"'], fill=fill(RED_BG), font=Font(name=FONT,size=10,bold=True,color=RED)))
ls.conditional_formatting.add(f"A{A}:H{B}", FormulaRule(
    formula=[f'AND($C{A}<>"",$C{A}={LATEST})'], fill=fill(INDIGO_LT), stopIfTrue=False))

# ════════════════════════════════════════════════════════════════════════════
# 4 · STOCK IN  — receipts and shrinkage
# ════════════════════════════════════════════════════════════════════════════
si = wb.create_sheet("Stock In")
widths(si, {"A":13,"B":14,"C":17,"D":14,"E":30,"F":8,"G":12,"H":13,"I":12})
si.column_dimensions["I"].hidden = True
shell(si, 9, [("A1","home"),("B1","items"),("C1","sell"),
              ("D1","stock"),("E1","inv"),("F1:G1","logs")], "A8", AMBER)
title(si,"A","H","GOODS RECEIVED","📥 Stock In",
      "Deliveries go in as IN. Damage, loss and theft go in as SHRINKAGE.")
header_row(si, 7, ["📅 Date","🔖 Ref / PO","🔎 SKU / Scan","🔄 Type","📦 Product",
                   "🔢 Qty","💲 Unit Cost","💵 Total","_SKU"], 38)
A2, B2 = SI_ROWS
heights(si, A2, B2, 32)
RECEIPTS = [("2026-08-01","PO-2001","LAV-001","IN",100),("2026-08-01","PO-2001","CER-014","IN",150),
            ("2026-08-01","PO-2002","TOT-007","IN",120),("2026-08-02","PO-2002","JRN-002","IN",80),
            ("2026-08-02","PO-2003","SKN-021","IN",60),("2026-08-02","PO-2003","TEA-009","IN",70),
            ("2026-08-11","ADJ-3001","TEA-009","SHRINKAGE",3)]
for r in range(A2, B2 + 1):
    i = r - A2
    if i < len(RECEIPTS):
        d, ref, sku, ty, q = RECEIPTS[i]
        si.cell(r,1,_dt.date(*[int(x) for x in d.split("-")]))
        si.cell(r,2,ref); si.cell(r,3,sku); si.cell(r,4,ty); si.cell(r,6,q)
    put(si,f"I{r}",f'=IF($C{r}="","",IF(COUNTIF({i_sku},$C{r})>0,$C{r},{XL(f"$C{r}",i_bar,i_sku)}))',
        fnt(9,False,MUTED),CARD,al("center","center"),bd=BOX)
    put(si,f"E{r}",f'=IF($C{r}="","",IF($I{r}="","⚠️ Unknown code",{XL(f"$I{r}",i_sku,i_name)}))',
        fnt(10,False,INK),CARD,al("left","center",ind=1),bd=BOX)
    put(si,f"G{r}",f'=IF($I{r}="","",{XL(f"$I{r}",i_sku,i_cost,"0")})',
        fnt(10,False,MUTED),CARD,al("center","center"),MONEY,BOX)
    put(si,f"H{r}",f'=IF($I{r}="","",$F{r}*$G{r})',fnt(11,True,INK),CARD,al("center","center"),MONEY,BOX)
    for c,(a_,f_) in {1:("center",DATE),2:("center",None),3:("center",None),4:("center",None),6:("center",INT)}.items():
        cell = si.cell(r,c); cell.fill = fill(INPUT_BG); cell.border = BOX
        cell.font = fnt(11, c in (2,3,4), INK); cell.alignment = al(a_,"center")
        if f_: cell.number_format = f_
dv4 = DataValidation(type="list", formula1="SKU_List", allow_blank=True,
                     showDropDown=False, showErrorMessage=False)
dv4.prompt = "Pick, type the first letters, or scan the barcode."; dv4.promptTitle = "🔎 Scan or select"
si.add_data_validation(dv4); dv4.add(f"C{A2}:C{B2}")
dv5 = DataValidation(type="list", formula1='"IN,SHRINKAGE"', allow_blank=True, showDropDown=False)
dv5.prompt = "IN = received from supplier · SHRINKAGE = damaged, lost or stolen"
dv5.promptTitle = "🔄 Movement type"
dv5.error = "Choose IN or SHRINKAGE."; dv5.errorTitle = "Invalid type"
si.add_data_validation(dv5); dv5.add(f"D{A2}:D{B2}")
si.conditional_formatting.add(f"D{A2}:D{B2}", FormulaRule(
    formula=[f'$D{A2}="IN"'], fill=fill(GREEN_BG), font=Font(name=FONT,size=11,bold=True,color=GREEN)))
si.conditional_formatting.add(f"D{A2}:D{B2}", FormulaRule(
    formula=[f'$D{A2}="SHRINKAGE"'], fill=fill(AMBER_BG), font=Font(name=FONT,size=11,bold=True,color=AMBER)))
si.conditional_formatting.add(f"E{A2}:E{B2}", FormulaRule(
    formula=[f'$E{A2}="⚠️ Unknown code"'], fill=fill(RED_BG), font=Font(name=FONT,size=10,bold=True,color=RED)))

# ════════════════════════════════════════════════════════════════════════════
# 5 · SALES LOG  — header view, one row per invoice, fully automatic
# ════════════════════════════════════════════════════════════════════════════
sl = wb.create_sheet("Sales Log")
widths(sl, {"A":14,"B":13,"C":22,"D":16,"E":24,"F":9,"G":11,"H":13,"I":11,"J":13,"K":14})
sl.column_dimensions["K"].hidden = True
shell(sl, 11, [("A1","home"),("B1:C1","items"),("D1","sell"),
               ("E1","stock"),("F1:G1","inv"),("H1:I1","logs")], "A8", INDIGO_DK)
title(sl,"A","J","HEADER LOG","📊 Sales Log",
      "One row per invoice — written automatically the moment a sale is entered on the Sell screen.")
header_row(sl, 7, ["🧾 Invoice No","📅 Date","👤 Customer","📞 Phone","✉️ Email",
                   "📦 Items","🔢 Qty","💵 Subtotal","🧾 Tax","💰 Amount","_pick"], 38)
A3_, B3_ = SL_ROWS
heights(sl, A3_, B3_, 32)
put(sl, f"K{A3_-1}", "LATEST", fnt(9,False,MUTED), CARD, al("center","center"))
NUM = f'{PREFIX}&TEXT(1000+ROW()-{A3_-1},"0000")'
for r in range(A3_, B3_ + 1):
    put(sl,f"A{r}",f'=IF(COUNTIF({l_inv},{NUM})=0,"",{NUM})',
        fnt(11,True,INDIGO_DK),CARD,al("center","center"),bd=BOX)
    put(sl,f"B{r}",f'=IF($A{r}="","",_xlfn.MAXIFS({l_date},{l_inv},$A{r}))',
        fnt(10,False,INK),CARD,al("center","center"),DATE,BOX)
    put(sl,f"F{r}",f'=IF($A{r}="","",COUNTIF({l_inv},$A{r}))',
        fnt(10,False,MUTED),CARD,al("center","center"),INT,BOX)
    put(sl,f"G{r}",f'=IF($A{r}="","",SUMIF({l_inv},$A{r},{l_qty}))',
        fnt(11,True,INK),CARD,al("center","center"),INT,BOX)
    put(sl,f"H{r}",f'=IF($A{r}="","",SUMIF({l_inv},$A{r},{l_tot}))',
        fnt(10,False,INK),CARD,al("center","center"),MONEY,BOX)
    put(sl,f"I{r}",f'=IF($A{r}="","",$H{r}*{TAXRATE})',
        fnt(10,False,MUTED),CARD,al("center","center"),MONEY,BOX)
    put(sl,f"J{r}",f'=IF($A{r}="","",$H{r}+$I{r})',
        fnt(12,True,GREEN),CARD,al("center","center"),MONEY,BOX)
    put(sl,f"K{r}",f'=IF($A{r}="","",$A{r})',fnt(9,False,MUTED),CARD,al("center","center"))
    for c in (3,4,5):
        cell = sl.cell(r,c); cell.fill = fill(INPUT_BG); cell.border = BOX
        cell.font = fnt(10,False,INK); cell.alignment = al("left","center",ind=1)
        if c == 4: cell.number_format = "@"
sl["C7"].comment = Comment(
    "Customer, phone and email are the only cells you fill here — and only if you want the "
    "one-tap WhatsApp / Email invoice buttons to know where to send.\n"
    "Everything else is written automatically from the Sell screen.", "Momentum Mind Store")
sl.conditional_formatting.add(f"A{A3_}:J{B3_}", FormulaRule(
    formula=[f'AND($A{A3_}<>"",$B{A3_}=TODAY())'], fill=fill(GREEN_BG), stopIfTrue=False))

# ════════════════════════════════════════════════════════════════════════════
# 6 · INVOICE GENERATOR  — opens on your latest sale, zero taps
# ════════════════════════════════════════════════════════════════════════════
iv = wb.create_sheet("Invoice Generator")
widths(iv, {"A":15,"B":28,"C":10,"D":15,"E":15,"F":14,"G":14,"H":40,"I":60})
for c in "HI": iv.column_dimensions[c].hidden = True
shell(iv, 9, [("A1","home"),("B1","items"),("C1:D1","sell"),
              ("E1","stock"),("F1","inv"),("G1","logs")], "A4", AMBER)
iv.merge_cells("A4:C4"); iv.merge_cells("A5:C5"); iv.merge_cells("D4:E5")
put(iv,"A4","=Settings!$C$8",fnt(20,True,INK),BG,al("left","center"))
put(iv,"A5",'=Settings!$C$9&"  ·  "&Settings!$C$10',fnt(10,False,MUTED),BG,al("left","center"))
put(iv,"D4","🧾 INVOICE",fnt(26,True,INDIGO),BG,al("right","center"))
iv.row_dimensions[4].height = 30; iv.row_dimensions[5].height = 24

put(iv,"A6","🔎 SHOW INVOICE",fnt(10,True,WHITE),INK,al("center","center"),bd=BOX)
iv.merge_cells("B6:C6")
put(iv,"B6","LATEST",fnt(12,True,INK),INPUT_BG,al("center","center"),bd=gold())
iv.merge_cells("D6:E6")
put(iv,"D6","Leave on LATEST and this screen always shows the sale you just made.",
    fnt(9,False,MUTED,i=True),BG,al("right","center"))
iv.row_dimensions[6].height = 34

put(iv,"A7","🧾 INVOICE NO",fnt(10,True,WHITE),INK,al("center","center"),bd=BOX)
iv.merge_cells("B7:C7")
put(iv,"B7",f'=IF($B$6="LATEST",{LATEST},$B$6)',fnt(15,True,INDIGO_DK),CARD,al("center","center"),bd=BOX)
put(iv,"D7","📅 Date",fnt(10,True,SLATE),BG,al("right","center"))
put(iv,"E7",f'=IFERROR({XL("$B$7",g_inv,g_date,"TODAY()")},TODAY())',fnt(11,False,INK),CARD,al("center","center"),DATE,BOX)
iv.row_dimensions[7].height = 38

put(iv,"A8","👤 BILL TO",fnt(10,True,WHITE),INK,al("center","center"),bd=BOX)
iv.merge_cells("B8:C8")
put(iv,"B8",f'=IFERROR({XL("$B$7",g_inv,g_cust)},"")',fnt(11,True,INK),CARD,al("left","center",ind=1),bd=BOX)
put(iv,"D8","Due",fnt(10,True,SLATE),BG,al("right","center"))
put(iv,"E8",'=IF($E$7="","",$E$7+Settings!$C$15)',fnt(11,False,INK),CARD,al("center","center"),DATE,BOX)
iv.row_dimensions[8].height = 32

put(iv,"A9","📞 / ✉️",fnt(10,True,WHITE),INK,al("center","center"),bd=BOX)
iv.merge_cells("B9:C9")
put(iv,"B9",f'=IFERROR({XL("$B$7",g_inv,g_phone)},"")',fnt(10,False,MUTED),CARD,al("left","center",ind=1),"@",BOX)
iv.merge_cells("D9:E9")
put(iv,"D9",f'=IFERROR({XL("$B$7",g_inv,g_mail)},"")',fnt(10,False,MUTED),CARD,al("left","center",ind=1),bd=BOX)
iv.row_dimensions[9].height = 30; iv.row_dimensions[10].height = 12

for ref,h in [("A11","SKU"),("B11","📦 Description"),("C11","Qty"),("D11","Unit Price"),("E11","Line Total")]:
    put(iv,ref,h,fnt(10,True,WHITE),INK,al("center","center"),bd=BOX)
iv.row_dimensions[11].height = 34
IVA, IVB = 12, 26
heights(iv, IVA, IVB, 30)
for r in range(IVA, IVB + 1):
    key = f'$B$7&"|"&{r-IVA+1}'
    put(iv,f"A{r}",f'=IFERROR(INDEX({l_sku},MATCH({key},{l_key},0)),"")',
        fnt(10,True,SLATE),CARD,al("center","center"),bd=BOX)
    put(iv,f"B{r}",f'=IF($A{r}="","",{XL(f"$A{r}",i_sku,i_name)})',
        fnt(10,False,INK),CARD,al("left","center",ind=1),bd=BOX)
    put(iv,f"C{r}",f'=IFERROR(INDEX({l_qty},MATCH({key},{l_key},0)),"")',
        fnt(10,False,INK),CARD,al("center","center"),INT,BOX)
    put(iv,f"D{r}",f'=IF($A{r}="","",{XL(f"$A{r}",i_sku,i_price,"0")})',
        fnt(10,False,INK),CARD,al("center","center"),MONEY,BOX)
    put(iv,f"E{r}",f'=IF($A{r}="","",$C{r}*$D{r})',
        fnt(10,True,INK),CARD,al("center","center"),MONEY,BOX)
    put(iv,f"H{r}",f'=IF($A{r}="","",$B{r}&"  x"&TEXT($C{r},"0")&"  "&TEXT($E{r},"0.00")&"%0A")',
        fnt(9,False,MUTED),BG,al("left","center"))

iv.row_dimensions[27].height = 10
for r,label,val,nf,bgc,txt,sz in [
        (28,"Subtotal",f"=SUM($E${IVA}:$E${IVB})",MONEY,CARD,INK,11),
        (29,"Tax rate",f"={TAXRATE}",PCT,CARD,MUTED,11),
        (30,"Tax amount","=$E$28*$E$29",MONEY,CARD,INK,11),
        (31,"Discount",0,MONEY,INPUT_BG,INK,11),
        (32,"GRAND TOTAL","=$E$28+$E$30-$E$31",MONEY,INDIGO,WHITE,15)]:
    iv.row_dimensions[r].height = 42 if r == 32 else 34
    iv.merge_cells(f"C{r}:D{r}")
    put(iv,f"C{r}",label,fnt(sz-1,True,txt if r==32 else SLATE),bgc if r==32 else BG,al("right","center"))
    put(iv,f"E{r}",val,fnt(sz,True,txt),bgc,al("center","center"),nf,gold() if r==31 else BOX)
    if r == 32:
        for c in (3,4,5): iv.cell(r,c).fill = fill(INDIGO); iv.cell(r,c).border = BOX

LINES = "&".join(f"$H${r}" for r in range(IVA, IVB + 1))
MSG = ('="*"&Settings!$C$8&"*%0AInvoice: "&$B$7&"%0ADate: "&TEXT($E$7,"dd-mmm-yyyy")'
       '&"%0A- - - - - - - - - -%0A"&' + LINES +
       '&"- - - - - - - - - -%0ASubtotal: "&TEXT($E$28,"0.00")&"%0ATax: "&TEXT($E$30,"0.00")'
       '&"%0A*TOTAL: "&TEXT($E$32,"0.00")&"*%0A%0A"&Settings!$C$16&"%0AThank you!"')
put(iv,"I7",MSG,fnt(9,False,MUTED),BG,al("left","center"))
put(iv,"I8",'=SUBSTITUTE(SUBSTITUTE($I$7,"%0A","%0D%0A"),"*","")',fnt(9,False,MUTED),BG,al("left","center"))
# wa.me accepts digits only - strip + - space ( ) . from the customer phone
PHONE_CLEAN = ('SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE('
               '$B$9,"+",""),"-","")," ",""),"(",""),")",""),".","")')
iv.row_dimensions[33].height = 12
iv.row_dimensions[34].height = 46
iv.merge_cells("A34:B34"); iv.merge_cells("C34:D34")
put(iv,"A34","📄  SAVE AS PDF   ›   File ▸ Export ▸ PDF",fnt(10,True,WHITE),INDIGO_DK,al("center","center",True),bd=BOX)
iv["A34"].comment = Comment(
    "Desktop: File ▸ Export ▸ Create PDF/XPS.\nPhone: ⋯ ▸ Export ▸ PDF.\n"
    "The print area is already set to exactly this invoice.\n\n"
    "For a true one-tap PDF button, import the PRO EDITION macro "
    "(see VBA_PRO_EDITION.bas in your download).", "Momentum Mind Store")
put(iv,"C34",f'=IF($B$7="","",HYPERLINK("https://wa.me/"&{PHONE_CLEAN}&"?text="&SUBSTITUTE($I$7," ","%20"),"💬  SEND ON WHATSAPP"))',
    fnt(11,True,WHITE),WA_GREEN,al("center","center"),bd=BOX)
put(iv,"E34",f'=IF($B$7="","",HYPERLINK("mailto:"&$D$9&"?subject="&SUBSTITUTE("Invoice "&$B$7&" from "&Settings!$C$8," ","%20")&"&body="&SUBSTITUTE($I$8," ","%20"),"✉️  EMAIL"))',
    fnt(11,True,WHITE),INDIGO,al("center","center"),bd=BOX)
for c in (1,2): iv.cell(34,c).fill = fill(INDIGO_DK); iv.cell(34,c).border = BOX
for c in (3,4): iv.cell(34,c).fill = fill(WA_GREEN); iv.cell(34,c).border = BOX

for r,f_,fo in [(36,'="Payment terms: "&Settings!$C$16',fnt(10,True,SLATE)),
                (37,"=Settings!$C$17",fnt(10,False,MUTED)),
                (38,'="Thank you for supporting "&Settings!$C$8&" 💛"',fnt(11,True,INDIGO))]:
    iv.merge_cells(f"A{r}:E{r}"); iv.row_dimensions[r].height = 26 if r < 38 else 34
    put(iv,f"A{r}",f_,fo,BG,al("center" if r==38 else "left","center"))

dv6 = DataValidation(type="list", formula1="INV_PICK", allow_blank=True, showDropDown=False)
dv6.prompt = "LATEST = always show the sale you just made. Or pick any past invoice number."
dv6.promptTitle = "🔎 Which invoice?"
iv.add_data_validation(dv6); dv6.add("B6")
iv.print_area = "A4:E38"
iv.page_setup.orientation = "portrait"
iv.page_setup.fitToWidth = 1; iv.page_setup.fitToHeight = 1
iv.sheet_properties.pageSetUpPr.fitToPage = True
iv.page_margins.left = iv.page_margins.right = 0.4
iv.page_margins.top = iv.page_margins.bottom = 0.5

# ════════════════════════════════════════════════════════════════════════════
# 7 · DASHBOARD
# ════════════════════════════════════════════════════════════════════════════
db = wb.create_sheet("Dashboard")
widths(db, {"A":14,"B":30,"C":11,"D":12,"E":14,"F":20,"G":18,"H":16,"I":14})
db.column_dimensions["I"].hidden = True
shell(db, 9, [("A1","home"),("B1","items"),("C1:D1","sell"),
              ("E1","stock"),("F1","inv"),("G1:H1","logs")], "A10", INDIGO)
title(db,"A","H","LIVE OVERVIEW","🏠 Dashboard",
      "Every number recalculates the moment you scan an item on the Sell screen.")
DA, DB_ = 10, 59
KPI1 = [("A7:B7","🏷️  ACTIVE SKUs",f"=COUNTA({i_sku})",INT,INDIGO_LT,INDIGO_DK),
        ("C7:D7","📦  UNITS IN STOCK",f"=SUM($E${DA}:$E${DB_})",INT,INDIGO_LT,INDIGO_DK),
        ("E7:F7","💵  TODAY'S SALES",f"=SUMIF({g_date},TODAY(),{g_amt})",MONEY,GREEN_BG,GREEN),
        ("G7:H7","⚠️  NEEDS REORDER",f'=COUNTIF($F${DA}:$F${DB_},"⚠️ REORDER NOW")',INT,RED_BG,RED)]
KPI2 = [("A8:B8","💰  STOCK VALUE (COST)",f"=SUM($I${DA}:$I${DB_})",MONEY,CARD,SLATE),
        ("C8:D8","🧾  ALL-TIME SALES",f"=SUM({g_amt})",MONEY,CARD,SLATE)]
for row, group in ((7, KPI1), (8, KPI2)):
    db.row_dimensions[row].height = 46 if row == 7 else 38
    for rng, label, f_, nf, bgc, txt in group:
        db.merge_cells(rng); a = rng.split(":")[0]
        put(db,a,f_,fnt(16 if row==7 else 13,True,txt),bgc,al("center","center"),nf,BOX)
        db[a].comment = Comment(label, "Momentum Mind Store")
        for c in range(openpyxl.utils.column_index_from_string(a[0]),
                       openpyxl.utils.column_index_from_string(rng.split(":")[1][0])+1):
            db.cell(row,c).fill = fill(bgc); db.cell(row,c).border = BOX
for ref, f_, bgc in [("E8",'=HYPERLINK("#\'Start Here\'!A1","🚀  Start Here")',SLATE),
                     ("F8",'=HYPERLINK("#\'Settings\'!A1","⚙️  Settings")',SLATE)]:
    put(db,ref,f_,fnt(10,True,WHITE),bgc,al("center","center"),bd=BOX)
db.merge_cells("G8:H8")
put(db,"G8",'=HYPERLINK("#\'Log Sale\'!A1","🛒  START A NEW SALE")',fnt(12,True,WHITE),GREEN,al("center","center"),bd=BOX)
for c in (7,8): db.cell(8,c).fill = fill(GREEN); db.cell(8,c).border = BOX

header_row(db, 9, ["SKU","📦 Product","⬇ Total IN","⬆ Total OUT","📊 Current Stock",
                   "🚦 Stock Status","💬 WhatsApp Order","✉️ Email Order","Value"], 36)
heights(db, DA, DB_, 32)
WA_TXT = ('"Hello "&{supp}&", we urgently need more "&$B{r}&". Current stock is only "'
          '&TEXT($E{r},"0")&" units. Please confirm availability and your fastest lead time. '
          'Thank you - Momentum Mind Store."')
ML_TXT = ('"Hello "&{supp}&"," & CHAR(10) & "We urgently need to restock "&$B{r}&" (SKU "&$A{r}&'
          '"). Our current stock is "&TEXT($E{r},"0")&" units, which is below our reorder threshold."'
          ' & CHAR(10) & "Please confirm availability, unit price and lead time by return." & CHAR(10)'
          ' & "Kind regards," & CHAR(10) & Settings!$C$8')
for r in range(DA, DB_ + 1):
    n = r - DA + 1
    supp = XL(f"$A{r}", i_sku, i_supp, '"Supplier"')
    ph = ('SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE('
          + XL(f"$A{r}", i_sku, i_phone) + ',"+",""),"-","")," ",""),"(",""),")",""),".","")')
    put(db,f"A{r}",f'=IF(INDEX({i_sku},{n})="","",INDEX({i_sku},{n}))',
        fnt(10,True,SLATE),CARD,al("left","center",ind=1),bd=BOX)
    put(db,f"B{r}",f'=IF($A{r}="","",{XL(f"$A{r}",i_sku,i_name)})',
        fnt(10,False,INK),CARD,al("left","center",ind=1),bd=BOX)
    put(db,f"C{r}",f'=IF($A{r}="","",SUMIFS({s_qty},{s_sku},$A{r},{s_type},"IN"))',
        fnt(10,False,MUTED),CARD,al("center","center"),INT,BOX)
    put(db,f"D{r}",f'=IF($A{r}="","",SUMIF({l_sku},$A{r},{l_qty})+SUMIFS({s_qty},{s_sku},$A{r},{s_type},"SHRINKAGE"))',
        fnt(10,False,MUTED),CARD,al("center","center"),INT,BOX)
    put(db,f"E{r}",f'=IF($A{r}="","",$C{r}-$D{r})',fnt(12,True,INK),CARD,al("center","center"),INT,BOX)
    put(db,f"F{r}",f'=IF($A{r}="","",IF($E{r}<={XL(f"$A{r}",i_sku,i_max,"0")}*0.1,"⚠️ REORDER NOW","✅ IN STOCK"))',
        fnt(10,True,INK),CARD,al("center","center"),bd=BOX)
    put(db,f"G{r}",f'=IF($A{r}="","",IF($F{r}="⚠️ REORDER NOW",HYPERLINK("https://wa.me/"&{ph}&"?text="'
                    f'&SUBSTITUTE({WA_TXT.format(supp=supp,r=r)}," ","%20"),"💬  ORDER NOW"),""))',
        fnt(10,True,WHITE),CARD,al("center","center"),bd=BOX)
    put(db,f"H{r}",f'=IF($A{r}="","",IF($F{r}="⚠️ REORDER NOW",HYPERLINK("mailto:"&{XL(f"$A{r}",i_sku,i_mail)}'
                    f'&"?subject="&SUBSTITUTE("URGENT restock request - "&$B{r}&" ("&$A{r}&")"," ","%20")'
                    f'&"&body="&SUBSTITUTE(SUBSTITUTE({ML_TXT.format(supp=supp,r=r)}," ","%20"),CHAR(10),"%0D%0A"),"✉️  EMAIL"),""))',
        fnt(10,True,WHITE),CARD,al("center","center"),bd=BOX)
    put(db,f"I{r}",f'=IF($A{r}="","",$E{r}*{XL(f"$A{r}",i_sku,i_cost,"0")})',
        fnt(10,False,MUTED),CARD,al("center","center"),MONEY,BOX)
for rng, f_, fl, fo in [
        (f"F{DA}:F{DB_}", f'$F{DA}="⚠️ REORDER NOW"', RED_BG, Font(name=FONT,size=10,bold=True,color=RED)),
        (f"F{DA}:F{DB_}", f'$F{DA}="✅ IN STOCK"', GREEN_BG, Font(name=FONT,size=10,bold=True,color=GREEN)),
        (f"E{DA}:E{DB_}", f'AND($A{DA}<>"",$E{DA}<=0)', RED_BG, Font(name=FONT,size=12,bold=True,color=RED)),
        (f"G{DA}:G{DB_}", f'$G{DA}<>""', WA_GREEN, Font(name=FONT,size=10,bold=True,color=WHITE)),
        (f"H{DA}:H{DB_}", f'$H{DA}<>""', INDIGO, Font(name=FONT,size=10,bold=True,color=WHITE))]:
    db.conditional_formatting.add(rng, FormulaRule(formula=[f_], fill=fill(fl), font=fo))

# ════════════════════════════════════════════════════════════════════════════
# 8 · SETTINGS
# ════════════════════════════════════════════════════════════════════════════
st = wb.create_sheet("Settings")
widths(st, {"A":3,"B":24,"C":20,"D":18,"E":18,"F":18,"G":18,"H":3})
shell(st, 8, [("A1:B1","home"),("C1","items"),("D1","sell"),
              ("E1","stock"),("F1","inv"),("G1:H1","logs")], "A4", SLATE)
title(st,"B","G","CONFIGURATION","⚙️ Settings","Fill these once. Every screen reads from here.")
SETTINGS = [(8,"🏪 Business name","Momentum Mind Store",None),
            (9,"✉️ Business email","hello@momentummindstore.com",None),
            (10,"📞 Business phone","+1 555 010 2030","@"),
            (11,"🌐 Website / Etsy","etsy.com/shop/MomentumMindStore",None),
            (12,"🧾 Tax rate",0.10,PCT),
            (13,"💱 Currency symbol","$",None),
            (14,"#️⃣ Invoice prefix","INV-",None),
            (15,"📆 Payment due (days)",14,INT),
            (16,"💳 Payment terms","Payment due within 14 days of invoice date.",None),
            (17,"🏦 Payment details","Bank: Example Bank · Acct: 0000 0000 · Ref: your invoice no.",None)]
for r,label,val,nf in SETTINGS:
    st.row_dimensions[r].height = 34
    put(st,f"B{r}",label,fnt(10,True,INK),CARD,al("left","center",ind=1),bd=BOX)
    st.merge_cells(f"C{r}:G{r}")
    put(st,f"C{r}",val,fnt(11,False,INK),INPUT_BG,al("left","center",ind=1),nf,gold())
    for c in range(4,8): st.cell(r,c).fill = fill(INPUT_BG); st.cell(r,c).border = gold()
st["B14"].comment = Comment(
    "Invoice numbers are built as prefix + 1001, 1002, 1003 … and are generated automatically "
    "each time you tick 🆕 on the Sell screen. Change the prefix here and every invoice renumbers.",
    "Momentum Mind Store")
for r,txt in [(19,"🔒 Rows 8–17 are the only cells you edit here. Everything else is automatic."),
              (20,"💱 The currency symbol above is documentation only — Excel number formats are static. "
                  "To change currency, select the money columns and apply e.g. £#,##0.00")]:
    st.merge_cells(f"B{r}:G{r}"); st.row_dimensions[r].height = 30
    put(st,f"B{r}",txt,fnt(9,False,MUTED,i=True),BG,al("left","center"))

# ── NAMES, ORDER, SAVE ──────────────────────────────────────────────────────
wb.defined_names.add(DefinedName("SKU_List", attr_text=f"{IM}!$A$4:$A$53"))
wb.defined_names.add(DefinedName("INV_PICK", attr_text=f"{SL}!$K$7:$K$107"))
ORDER = ["Start Here","Dashboard","Item Master","Log Sale","Stock In",
         "Sales Log","Invoice Generator","Settings"]
wb._sheets = [wb[n] for n in ORDER]
wb.active = 0
for ws in wb.worksheets:
    ws.sheet_view.zoomScale = 100
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False
    ws.sheet_view.tabSelected = (ws.title == "Start Here")

OUT = "Momentum-Mind-Inventory-Manager-PRO.xlsx"
wb.save(OUT)
print("Built:", OUT)
print("Sheets:", wb.sheetnames)
